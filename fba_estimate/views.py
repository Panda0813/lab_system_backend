from django.db import connection, close_old_connections, transaction
from django.db.models import Q, F, Value, CharField, IntegerField
from rest_framework import generics, serializers
from rest_framework import filters, status
from rest_framework.decorators import api_view
from rest_framework.response import Response

from fba_estimate.serializers import FirstServiceSerializer, SecondServiceSerializer, CompanySerializer, \
    EstimateOptionSerializer, CapitalSurplusSerializer
from fba_estimate.serializers import EstimateMonthDetailSerializer, EstimateMonthFutureSerializer
from fba_estimate.models import FirstService, SecondService, Company, EstimateOption, CapitalSurplus
from fba_estimate.models import EstimateMonthDetail, EstimateMonthFuture
from equipments.ext_utils import REST_FAIL, REST_SUCCESS, create_excel_resp, dictfetchall
from utils.log_utils import set_create_log, set_update_log, set_delete_log
from fba_estimate.ext_utils import get_last_month, get_file_path
from dateutil.relativedelta import relativedelta
from decimal import Decimal
from openpyxl import load_workbook
from copy import copy

import os
import calendar
import traceback
import time
import datetime
import pandas as pd
import logging

logger = logging.getLogger('django')


@api_view(['GET'])
def get_service_tree(request):
    first_obj = FirstService.objects.filter(is_delete=False)
    if not first_obj:
        return REST_SUCCESS([])
    first_qs = list(first_obj.annotate(type=Value('first', output_field=CharField())).extra(select={'label': 'name'}).
                    values('id', 'label', 'is_active', 'type', 'create_time'))
    for item in first_qs:
        second_obj = SecondService.objects.filter(first_service_id=item['id'], is_delete=False)
        if second_obj:
            item['children'] = list(second_obj.annotate(type=Value('second', output_field=CharField()),
                                                        first_service_name=Value(item['label'], output_field=CharField())).
                                    extra(select={'label': 'name'}).
                                    values('id', 'label', 'is_active', 'type', 'first_service_name', 'create_time'))
        else:
            item['children'] = []
    for f in first_qs:
        first_service = f['id']
        f['id'] = 'first_' + str(f['id'])
        f['create_time'] = f['create_time'].strftime('%Y-%m-%d %H:%M:%S')
        for s in f['children']:
            s['id'] = 'second_' + str(s['id'])
            s['first_service'] = first_service
            s['create_time'] = s['create_time'].strftime('%Y-%m-%d %H:%M:%S')
    return REST_SUCCESS(first_qs)


@api_view(['GET'])
def get_service_select(request):
    first_obj = FirstService.objects.filter(is_delete=False, is_active=True)
    if not first_obj:
        return REST_SUCCESS([])
    first_qs = list(first_obj.extra(select={'value': 'id', 'label': 'name'}).values('value', 'label'))
    for item in first_qs:
        second_obj = SecondService.objects.filter(first_service_id=item['value'], is_delete=False, is_active=True)
        if second_obj:
            item['children'] = list(second_obj.extra(select={'value': 'id', 'label': 'name'}).values('value', 'label'))
        else:
            item['children'] = []
    return REST_SUCCESS(first_qs)


class FirstServiceList(generics.ListCreateAPIView):
    model = FirstService
    queryset = model.objects.filter(is_delete=False).all()
    serializer_class = FirstServiceSerializer
    table_name = model._meta.db_table
    verbose_name = model._meta.verbose_name

    @set_create_log
    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        name = request.data.get('name')
        exist_qs = FirstService.objects.filter(name=name)
        if exist_qs:
            serializer.is_valid(raise_exception=False)
            exist_qs.update(is_delete=False, create_time=datetime.datetime.now())
            new_data = {'id': exist_qs.first().id}
            new_data.update(serializer.data)
            headers = self.get_success_headers(serializer.data)
            return Response(new_data, status=status.HTTP_201_CREATED, headers=headers)
        else:
            serializer.is_valid(raise_exception=True)
            self.perform_create(serializer)
            headers = self.get_success_headers(serializer.data)
            return Response(serializer.data, status=status.HTTP_201_CREATED, headers=headers)


class FirstServiceDetail(generics.RetrieveUpdateDestroyAPIView):
    model = FirstService
    queryset = model.objects.filter(is_delete=False).all()
    serializer_class = FirstServiceSerializer
    table_name = model._meta.db_table
    verbose_name = model._meta.verbose_name

    @set_update_log
    def update(self, request, *args, **kwargs):
        partial = kwargs.pop('partial', False)
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data, partial=partial)
        serializer.is_valid(raise_exception=True)
        self.perform_update(serializer)

        if getattr(instance, '_prefetched_objects_cache', None):
            # If 'prefetch_related' has been applied to a queryset, we need to
            # forcibly invalidate the prefetch cache on the instance.
            instance._prefetched_objects_cache = {}

        return Response(serializer.data)

    @set_delete_log
    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        self.perform_destroy(instance)
        return REST_SUCCESS({'msg': '删除成功'})


class SecondServiceList(generics.ListCreateAPIView):
    model = SecondService
    queryset = model.objects.filter(is_delete=False).all()
    serializer_class = SecondServiceSerializer
    table_name = model._meta.db_table
    verbose_name = model._meta.verbose_name

    @set_create_log
    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        first_service = request.data.get('first_service')
        name = request.data.get('name')
        exist_qs = SecondService.objects.filter(name=name, first_service_id=first_service)
        if exist_qs:
            serializer.is_valid(raise_exception=False)
            exist_qs.update(is_delete=False, create_time=datetime.datetime.now())
            new_data = {'id': exist_qs.first().id}
            new_data.update(serializer.data)
            headers = self.get_success_headers(serializer.data)
            return Response(new_data, status=status.HTTP_201_CREATED, headers=headers)
        else:
            serializer.is_valid(raise_exception=True)
            self.perform_create(serializer)
            headers = self.get_success_headers(serializer.data)
            return Response(serializer.data, status=status.HTTP_201_CREATED, headers=headers)


class SecondServiceDetail(generics.RetrieveUpdateDestroyAPIView):
    model = SecondService
    queryset = model.objects.filter(is_delete=False).all()
    serializer_class = SecondServiceSerializer
    table_name = model._meta.db_table
    verbose_name = model._meta.verbose_name

    @set_update_log
    def update(self, request, *args, **kwargs):
        partial = kwargs.pop('partial', False)
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data, partial=partial)
        serializer.is_valid(raise_exception=True)
        self.perform_update(serializer)

        if getattr(instance, '_prefetched_objects_cache', None):
            # If 'prefetch_related' has been applied to a queryset, we need to
            # forcibly invalidate the prefetch cache on the instance.
            instance._prefetched_objects_cache = {}

        return Response(serializer.data)

    @set_delete_log
    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        self.perform_destroy(instance)
        return REST_SUCCESS({'msg': '删除成功'})


# 查询表中的收、付款公司
@api_view(['GET'])
def get_company(request):
    in_qs1 = EstimateMonthFuture.objects.values('in_company').distinct()
    in_qs2 = EstimateMonthDetail.objects.values('in_company').distinct()
    out_qs1 = EstimateMonthFuture.objects.values('out_company').distinct()
    out_qs2 = EstimateMonthDetail.objects.values('out_company').distinct()
    company_ls = {'in_company_ls': [], 'out_company_ls': []}
    in_qs = list(in_qs1) + list(in_qs2)
    out_qs = list(out_qs1) + list(out_qs2)
    if in_qs:
        in_company_ls = [item['in_company'] for item in in_qs if item['in_company']]
        company_ls['in_company_ls'] = list(set(in_company_ls))
    if out_qs:
        out_company_ls = [item['out_company'] for item in out_qs if item['out_company']]
        company_ls['out_company_ls'] = list(set(out_company_ls))
    return REST_SUCCESS(company_ls)


# 查询表中的填报人
@api_view(['GET'])
def get_writer_user(request):
    qs1 = EstimateMonthFuture.objects.values('writer_user').distinct()
    qs2 = EstimateMonthDetail.objects.values('writer_user').distinct()
    writer_user_ls = []
    qs = list(qs1) + list(qs2)
    if qs:
        writer_user_ls = [item['writer_user'] for item in qs if item['writer_user']]
        writer_user_ls = list(set(writer_user_ls))
    return REST_SUCCESS(writer_user_ls)


class CompanyList(generics.ListCreateAPIView):
    model = Company
    queryset = model.objects.filter(is_delete=False).all()
    serializer_class = CompanySerializer
    table_name = model._meta.db_table
    verbose_name = model._meta.verbose_name
    filter_backends = [filters.SearchFilter]
    search_fields = ['name']

    @set_create_log
    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        name = request.data.get('name')
        exist_qs = Company.objects.filter(name=name)
        if exist_qs:
            serializer.is_valid(raise_exception=False)
            exist_qs.update(is_delete=False, create_time=datetime.datetime.now())
            new_data = {'id': exist_qs.first().id}
            new_data.update(serializer.data)
            headers = self.get_success_headers(serializer.data)
            return Response(new_data, status=status.HTTP_201_CREATED, headers=headers)
        else:
            serializer.is_valid(raise_exception=True)
            self.perform_create(serializer)
            headers = self.get_success_headers(serializer.data)
            return Response(serializer.data, status=status.HTTP_201_CREATED, headers=headers)


class CompanyDetail(generics.RetrieveUpdateDestroyAPIView):
    model = Company
    queryset = model.objects.filter(is_delete=False).all()
    serializer_class = CompanySerializer
    table_name = model._meta.db_table
    verbose_name = model._meta.verbose_name

    @set_update_log
    def update(self, request, *args, **kwargs):
        partial = kwargs.pop('partial', False)
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data, partial=partial)
        serializer.is_valid(raise_exception=True)
        self.perform_update(serializer)

        if getattr(instance, '_prefetched_objects_cache', None):
            # If 'prefetch_related' has been applied to a queryset, we need to
            # forcibly invalidate the prefetch cache on the instance.
            instance._prefetched_objects_cache = {}

        return Response(serializer.data)

    @set_delete_log
    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        self.perform_destroy(instance)
        return REST_SUCCESS({'msg': '删除成功'})


class EstimateOptionList(generics.ListCreateAPIView):
    model = EstimateOption
    queryset = model.objects.all()
    serializer_class = EstimateOptionSerializer
    table_name = model._meta.db_table
    verbose_name = model._meta.verbose_name
    filter_backends = [filters.SearchFilter]
    search_fields = ['name']

    @set_create_log
    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        self.perform_create(serializer)
        headers = self.get_success_headers(serializer.data)
        return Response(serializer.data, status=status.HTTP_201_CREATED, headers=headers)


class EstimateOptionDetail(generics.RetrieveUpdateAPIView):
    model = EstimateOption
    queryset = model.objects.all()
    serializer_class = EstimateOptionSerializer
    table_name = model._meta.db_table
    verbose_name = model._meta.verbose_name

    @set_update_log
    def update(self, request, *args, **kwargs):
        partial = kwargs.pop('partial', False)
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data, partial=partial)
        serializer.is_valid(raise_exception=True)
        self.perform_update(serializer)

        if getattr(instance, '_prefetched_objects_cache', None):
            # If 'prefetch_related' has been applied to a queryset, we need to
            # forcibly invalidate the prefetch cache on the instance.
            instance._prefetched_objects_cache = {}

        return Response(serializer.data)


# 查询上月剩余资金
@api_view(['GET'])
def get_last_month_surplus(request):
    year, month = get_last_month()
    qs = CapitalSurplus.objects.filter(year=year, month=month)
    surplus = {'year': year, 'month': month}
    if qs:
        surplus = list(qs.values())[0]
    return REST_SUCCESS(surplus)


class CapitalSurplusList(generics.ListCreateAPIView):
    model = CapitalSurplus
    queryset = model.objects.all().order_by('-create_time')
    serializer_class = CapitalSurplusSerializer
    table_name = model._meta.db_table
    verbose_name = model._meta.verbose_name

    def list(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())
        year = request.GET.get('year')
        if year:
            queryset = queryset.filter(year=year)
        month = request.GET.get('month')
        if month:
            queryset = queryset.filter(month=month)

        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return self.get_paginated_response(serializer.data)

        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)

    @set_create_log
    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        self.perform_create(serializer)
        headers = self.get_success_headers(serializer.data)
        return Response(serializer.data, status=status.HTTP_201_CREATED, headers=headers)


class CapitalSurplusDetail(generics.RetrieveUpdateAPIView):
    model = CapitalSurplus
    queryset = model.objects.all()
    serializer_class = CapitalSurplusSerializer
    table_name = model._meta.db_table
    verbose_name = model._meta.verbose_name

    @set_update_log
    def update(self, request, *args, **kwargs):
        partial = kwargs.pop('partial', False)
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data, partial=partial)
        serializer.is_valid(raise_exception=True)
        self.perform_update(serializer)

        if getattr(instance, '_prefetched_objects_cache', None):
            # If 'prefetch_related' has been applied to a queryset, we need to
            # forcibly invalidate the prefetch cache on the instance.
            instance._prefetched_objects_cache = {}

        return Response(serializer.data)


def get_write_data_type(now_day):
    allow_option = EstimateOption.objects.all().first()
    if not allow_option:
        return None
    first_start_day = allow_option.first_start_day
    first_end_day = allow_option.first_end_day
    second_start_day = allow_option.second_start_day
    second_end_day = allow_option.second_end_day
    if first_start_day <= now_day <= first_end_day:
        return 1  # 预估数
    elif second_start_day <= now_day <= second_end_day:
        return 2  # 修正数
    else:
        return None


# 获取当日可填报的条件
@api_view(['GET'])
def get_write_option(request):
    now_date = datetime.datetime.now().date()
    now_day = now_date.day
    data_type = get_write_data_type(now_day)
    is_allow_write = True
    if not data_type:
        is_allow_write = False
    now_year = now_date.year
    now_month = now_date.month
    last_day = calendar.monthrange(now_year, now_month)[1]
    day_start = datetime.date(now_year, now_month, 1).strftime('%Y-%m-%d')
    day_end = datetime.date(now_year, now_month, last_day).strftime('%Y-%m-%d')
    month_ls = []
    for i in range(1, 6):
        t = now_date + relativedelta(months=i)
        month_ls.append(t.strftime('%Y-%m'))
    data = {
        'is_allow_write': is_allow_write,
        'data_type': data_type,
        'start_day': day_start,
        'end_day': day_end,
        'month_ls': month_ls,
        'start_month': month_ls[0],
        'end_month': month_ls[-1],
        'now_month': now_date.strftime('%Y-%m')
    }
    return REST_SUCCESS(data)


class MonthDetailList(generics.ListCreateAPIView):
    queryset = EstimateMonthDetail.objects.all().order_by('-create_time')
    serializer_class = EstimateMonthDetailSerializer

    def list(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())
        now_date = datetime.datetime.now()
        queryset = queryset.filter(create_time__year=now_date.year, create_time__month=now_date.month)
        req_user = request.user
        user_roles = req_user.role_set.values('id', 'role_code')
        user_roles = [item['role_code'] for item in list(user_roles)]
        if list(set(['developer', 'fbaManager']) & set(user_roles)):
            pass
        elif 'fbaCsManager' in user_roles:
            queryset = queryset.filter(first_service__name__in=['标准DRAM', '模组'])
        elif 'fbaSuprManager' in user_roles:
            queryset = queryset.filter(first_service__name__in=['大带宽产品', '设计服务'])
        else:
            queryset = queryset.filter(writer_user=req_user.username)
        # existTag = request.GET.get('existTag')
        # if existTag:  # 判断是否是查询已存在的填写
        # queryset = queryset.filter(writer_user=req_user.username)
        start_day = request.GET.get('start_time')
        end_day = request.GET.get('end_time')
        if start_day and end_day:
            queryset = queryset.filter(write_date__range=[start_day, end_day])
        in_company = request.GET.get('in_company')
        if in_company:
            queryset = queryset.filter(in_company__contains=in_company)
        out_company = request.GET.get('out_company')
        if out_company:
            queryset = queryset.filter(out_company__contains=out_company)
        money_type = request.GET.get('money_type')
        if money_type:
            queryset = queryset.filter(eval('~Q({}=None)'.format(money_type)))
        writer_user = request.GET.get('writer_user')
        if writer_user:
            writer_user_ls = writer_user.split(',')
            queryset = queryset.filter(writer_user__in=writer_user_ls)

        fuzzy_params = {}
        fuzzy_params['first_service_id'] = request.GET.get('first_service')
        fuzzy_params['second_service_id'] = request.GET.get('second_service')
        fuzzy_params['region'] = request.GET.get('region')
        fuzzy_params['data_type'] = request.GET.get('data_type')
        fuzzy_params['write_date'] = request.GET.get('write_date')
        fuzzy_params['year'] = request.GET.get('year')
        fuzzy_params['month'] = request.GET.get('month')
        fuzzy_params['day'] = request.GET.get('day')

        filter_params = {}
        for k, v in fuzzy_params.items():
            if v != None and v != '':
                filter_params[k] = v

        if filter_params:
            queryset = queryset.filter(**filter_params)

        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return self.get_paginated_response(serializer.data)

        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)

    def create(self, request, *args, **kwargs):
        now_day = datetime.datetime.now().day
        data_type = get_write_data_type(now_day)
        if not data_type:
            return REST_SUCCESS({'is_close_time': True})
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        data = serializer.validated_data.copy()
        writer_user = request.user.username
        data.update({'writer_user': writer_user})
        serializer.validated_data.update(data)
        self.perform_create(serializer)
        headers = self.get_success_headers(serializer.data)
        return Response(serializer.data, status=status.HTTP_201_CREATED, headers=headers)


class MonthDetailOperate(generics.RetrieveUpdateDestroyAPIView):
    model = EstimateMonthDetail
    queryset = model.objects.all()
    serializer_class = EstimateMonthDetailSerializer
    table_name = model._meta.db_table
    verbose_name = model._meta.verbose_name

    @set_update_log
    def update(self, request, *args, **kwargs):
        remarkTag = request.data.get('remarkTag')
        now_day = datetime.datetime.now().day
        data_type = get_write_data_type(now_day)
        if not (data_type or remarkTag):
            return REST_SUCCESS({'is_close_time': True})
        partial = kwargs.pop('partial', False)
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data, partial=partial)
        serializer.is_valid(raise_exception=True)
        self.perform_update(serializer)

        if getattr(instance, '_prefetched_objects_cache', None):
            # If 'prefetch_related' has been applied to a queryset, we need to
            # forcibly invalidate the prefetch cache on the instance.
            instance._prefetched_objects_cache = {}

        return Response(serializer.data)

    @set_delete_log
    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        self.perform_destroy(instance)
        return REST_SUCCESS({'msg': '删除成功'})


class MonthFutureList(generics.ListCreateAPIView):
    queryset = EstimateMonthFuture.objects.all().order_by('-create_time')
    serializer_class = EstimateMonthFutureSerializer

    def list(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())
        now_date = datetime.datetime.now()
        queryset = queryset.filter(create_time__year=now_date.year, create_time__month=now_date.month)
        req_user = request.user
        user_roles = req_user.role_set.values('id', 'role_code')
        user_roles = [item['role_code'] for item in list(user_roles)]
        if list(set(['developer', 'fbaManager']) & set(user_roles)):
            pass
        elif 'fbaCsManager' in user_roles:
            queryset = queryset.filter(first_service__name__in=['标准DRAM', '模组'])
        elif 'fbaSuprManager' in user_roles:
            queryset = queryset.filter(first_service__name__in=['大带宽产品', '设计服务'])
        else:
            queryset = queryset.filter(writer_user=req_user.username)
        # existTag = request.GET.get('existTag')
        # if existTag:  # 判断是否是查询已存在的填写
        # queryset = queryset.filter(writer_user=req_user.username)
        start_month = request.GET.get('start_time')
        end_month = request.GET.get('end_time')
        if start_month and end_month:
            queryset = queryset.filter(write_date__range=[start_month, end_month])
        in_company = request.GET.get('in_company')
        if in_company:
            queryset = queryset.filter(in_company__contains=in_company)
        out_company = request.GET.get('out_company')
        if out_company:
            queryset = queryset.filter(out_company__contains=out_company)
        money_type = request.GET.get('money_type')
        if money_type:
            queryset = queryset.filter(eval('~Q({}=None)'.format(money_type)))
        writer_user = request.GET.get('writer_user')
        if writer_user:
            writer_user_ls = writer_user.split(',')
            queryset = queryset.filter(writer_user__in=writer_user_ls)

        fuzzy_params = {}
        fuzzy_params['first_service_id'] = request.GET.get('first_service')
        fuzzy_params['second_service_id'] = request.GET.get('second_service')
        fuzzy_params['region'] = request.GET.get('region')
        fuzzy_params['data_type'] = request.GET.get('data_type')
        fuzzy_params['write_date'] = request.GET.get('write_date')
        fuzzy_params['year'] = request.GET.get('year')
        fuzzy_params['month'] = request.GET.get('month')

        filter_params = {}
        for k, v in fuzzy_params.items():
            if v != None and v != '':
                filter_params[k] = v

        if filter_params:
            queryset = queryset.filter(**filter_params)

        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return self.get_paginated_response(serializer.data)

        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)

    def create(self, request, *args, **kwargs):
        now_day = datetime.datetime.now().day
        data_type = get_write_data_type(now_day)
        if not data_type:
            return REST_SUCCESS({'is_close_time': True})
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        data = serializer.validated_data.copy()
        writer_user = request.user.username
        data.update({'writer_user': writer_user})
        serializer.validated_data.update(data)
        self.perform_create(serializer)
        headers = self.get_success_headers(serializer.data)
        return Response(serializer.data, status=status.HTTP_201_CREATED, headers=headers)


class MonthFutureOperate(generics.RetrieveUpdateDestroyAPIView):
    model = EstimateMonthFuture
    queryset = model.objects.all()
    serializer_class = EstimateMonthFutureSerializer
    table_name = model._meta.db_table
    verbose_name = model._meta.verbose_name

    @set_update_log
    def update(self, request, *args, **kwargs):
        remarkTag = request.data.get('remarkTag')
        now_day = datetime.datetime.now().day
        data_type = get_write_data_type(now_day)
        if not (data_type or remarkTag):
            return REST_SUCCESS({'is_close_time': True})
        partial = kwargs.pop('partial', False)
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data, partial=partial)
        serializer.is_valid(raise_exception=True)
        self.perform_update(serializer)

        if getattr(instance, '_prefetched_objects_cache', None):
            # If 'prefetch_related' has been applied to a queryset, we need to
            # forcibly invalidate the prefetch cache on the instance.
            instance._prefetched_objects_cache = {}

        return Response(serializer.data)

    @set_delete_log
    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        self.perform_destroy(instance)
        return REST_SUCCESS({'msg': '删除成功'})


def get_service_filter(req_user):
    user_roles = req_user.role_set.values('id', 'role_code')
    user_roles = [item['role_code'] for item in list(user_roles)]
    if list(set(['developer', 'fbaManager']) & set(user_roles)):
        return None
    elif 'fbaCsManager' in user_roles:
        return ['标准DRAM', '模组']
    elif 'fbaSuprManager' in user_roles:
        return ['大带宽产品', '设计服务']
    else:
        return 'fbaStandardUser'


def get_base_data(export_month, service_filter, username, export_type):
    now_date = datetime.datetime.now()
    year = now_date.year
    month = now_date.month
    qs1 = EstimateMonthDetail.objects
    qs2 = EstimateMonthFuture.objects
    if export_type == 'monitor':
        qs1 = qs1.filter(first_service__name__in=['标准DRAM', '模组', '大带宽产品', '设计服务'])
        qs2 = qs2.filter(first_service__name__in=['标准DRAM', '模组', '大带宽产品', '设计服务'])
    if isinstance(service_filter, list):
        qs1 = qs1.filter(first_service__name__in=service_filter)
        qs2 = qs2.filter(first_service__name__in=service_filter)
    elif service_filter == 'fbaStandardUser':
        qs1 = qs1.filter(writer_user=username)
        qs2 = qs2.filter(writer_user=username)
    if export_month != now_date.strftime('%Y-%m'):  # 导出时间非本月,直接导出修正数
        export_date = datetime.datetime.strptime(export_month, '%Y-%m')
        year = export_date.year
        month = export_date.month
        data_type = 2
        last_month = export_date + relativedelta(months=-1)
    else:  # 导出本月数据
        allow_option = EstimateOption.objects.all().first()
        second_start_day = allow_option.second_start_day
        if now_date.day < second_start_day:
            data_type = 1
        else:
            data_type = 2
        last_month = now_date + relativedelta(months=-1)
    qs1 = qs1.filter(year=year, month=month)
    qs2 = qs2.filter(create_time__year=year, create_time__month=month)
    if export_type == 'detail':
        qs1 = qs1.filter(data_type=data_type)
        qs2 = qs2.filter(data_type=data_type)
    value_args = ['first_service_id', 'first_service__name', 'second_service_id', 'second_service__name',
                  'region', 'in_company', 'out_company', 'year', 'month', 'in_usd', 'in_cny', 'out_usd', 'out_cny',
                  'writer_user', 'remarks']
    if export_type == 'monitor':
        value_args.append('data_type')
    qs2 = qs2.values(*tuple(value_args))
    value_args.append('day')
    qs1 = qs1.values(*tuple(value_args))
    surplus_qs = CapitalSurplus.objects.filter(year=last_month.year, month=last_month.month). \
        values('uniic_cny', 'uniic_usd', 'hk_usd', 'currency_exchange')  # 上月末资金余额
    surplus_data = {}
    if surplus_qs:
        surplus_data = list(surplus_qs)[0]
    return year, month, last_month, qs1, qs2, surplus_data


# 导出资金预估汇总表
@api_view(['GET'])
def export_month_detail(request):
    try:
        export_month = request.GET.get('exportMonth')
        if not export_month:
            return REST_FAIL({'导出时间不能为空'})
        req_user = request.user
        service_filter = get_service_filter(req_user)
        year, month, last_month, qs1, qs2, surplus_data = get_base_data(export_month, service_filter, req_user.username, 'detail')
        qs = list(qs1) + list(qs2)
        title = '{}年{}月份资金预估表'.format(year, month)
        template_path = os.path.dirname(__file__) + '/export_template/FBA_Month_Detail_Template.xlsx'
        if not qs:
            return create_excel_resp(template_path, title)
        df = pd.DataFrame(qs)
        if not qs1:
            df['day'] = None
        df.rename(columns={'first_service__name': 'first_service', 'second_service__name': 'second_service'}, inplace=True)
        ndf = df[['first_service_id', 'first_service', 'second_service_id', 'second_service',
                  'region', 'in_company', 'out_company', 'year',
                  'month', 'day', 'in_usd', 'in_cny', 'out_usd', 'out_cny', 'writer_user', 'remarks']]
        ndf[['in_usd', 'in_cny', 'out_usd', 'out_cny']] = ndf[['in_usd', 'in_cny', 'out_usd', 'out_cny']].replace({None: 0})
        ndf[['in_usd', 'in_cny', 'out_usd', 'out_cny']] = ndf[['in_usd', 'in_cny', 'out_usd', 'out_cny']].astype(float)
        region_type = {1: '中国南区', 2: '中国北区'}
        ndf['region'] = ndf['region'].map(region_type)
        ndf[['second_service', 'second_service_id', 'day', 'region', 'in_company', 'out_company']] = ndf[
            ['second_service', 'second_service_id', 'day', 'region', 'in_company', 'out_company']].replace({None: ''})
        ndf['remarks'] = ndf['remarks'].map(lambda x: [x] if x else [])
        remarkSum = ndf.groupby(['first_service_id', 'first_service', 'second_service_id', 'second_service',
                                 'region', 'in_company', 'out_company',
                                 'year', 'month', 'day', 'writer_user'])['remarks'].sum().reset_index()
        remarkSum['remarks'] = remarkSum['remarks'].map(lambda x: ','.join(x) if x else '')
        remarkSum.sort_values(by=['month', 'day', 'first_service_id', 'second_service_id'], inplace=True)
        remarkSum.drop(['first_service_id', 'second_service_id'], axis=1, inplace=True)
        moneySum = ndf.groupby(['first_service_id', 'first_service', 'second_service_id', 'second_service',
                                'region', 'in_company', 'out_company',
                                'year', 'month', 'day', 'writer_user']).sum().reset_index()
        moneySum.sort_values(by=['month', 'day', 'first_service_id', 'second_service_id'], inplace=True)
        moneySum.drop(['first_service_id', 'second_service_id'], axis=1, inplace=True)
        mergedf = pd.merge(moneySum, remarkSum,
                           on=['first_service', 'second_service', 'region', 'in_company', 'out_company',
                               'year', 'month', 'day', 'writer_user'],
                           how='outer')
        mergedf[['in_usd', 'in_cny', 'out_usd', 'out_cny']] = mergedf[
            ['in_usd', 'in_cny', 'out_usd', 'out_cny']].replace({0: None})
        wb = load_workbook(template_path)
        ws = wb['Sheet1']
        ws['B1'] = title
        ws['N1'] = float(surplus_data.get('uniic_usd')) if surplus_data.get('uniic_usd') else 0
        ws['N2'] = float(surplus_data.get('uniic_cny')) if surplus_data.get('uniic_cny') else 0
        ws['N3'] = float(surplus_data.get('hk_usd')) if surplus_data.get('hk_usd') else 0
        ws['O1'] = float(surplus_data.get('currency_exchange')) if surplus_data.get('currency_exchange') else 0
        ws['F6'] = year
        ws['G6'] = month
        ws['H6'] = 1
        insert_data = mergedf.to_dict('records')
        startRow = 7
        for i in range(len(insert_data)):
            data = insert_data[i]
            now_row = i + startRow
            ws.cell(row=now_row, column=1).value = data['first_service']
            ws.cell(row=now_row, column=2).value = data['second_service']
            ws.cell(row=now_row, column=3).value = data['region']
            ws.cell(row=now_row, column=4).value = data['in_company']
            ws.cell(row=now_row, column=5).value = data['out_company']
            ws.cell(row=now_row, column=6).value = data['year']
            ws.cell(row=now_row, column=7).value = data['month']
            ws.cell(row=now_row, column=8).value = data['day']
            ws.cell(row=now_row, column=9).value = data['in_usd']
            ws.cell(row=now_row, column=10).value = data['in_cny']
            ws.cell(row=now_row, column=11).value = data['out_usd']
            ws.cell(row=now_row, column=12).value = data['out_cny']
            ws.cell(row=now_row, column=13).value = "=M{}+I{}-K{}".format(now_row - 1, now_row, now_row)
            ws.cell(row=now_row, column=14).value = "=N{}+J{}-L{}".format(now_row - 1, now_row, now_row)
            ws.cell(row=now_row, column=15).value = "=M{}*$O$1+N{}".format(now_row, now_row)
            ws.cell(row=now_row, column=16).value = data['writer_user']
            ws.cell(row=now_row, column=17).value = data['remarks']
        for row in tuple(ws['A{}'.format(startRow + 1): 'Q{}'.format(len(insert_data) + startRow - 1)]):
            for cell in row:
                template_cell = ws.cell(row=startRow, column=cell.column)
                cell.border = copy(ws.cell(row=startRow, column=cell.column).border)
                cell.font = copy(template_cell.font)
                cell.fill = copy(template_cell.fill)
                cell.number_format = copy(template_cell.number_format)
                cell.alignment = copy(template_cell.alignment)
        save_path = get_file_path('detail', 'export_files')
        wb.save(save_path)
        return create_excel_resp(save_path, title)
    except Exception as e:
        logger.error('导出失败, error: {}'.format(traceback.format_exc()))
        return REST_FAIL({'msg': '导出失败, error: {}'.format(str(e))})


def create_month_ls(export_month):
    export_date = datetime.datetime.strptime(export_month, '%Y-%m')
    month_ls = []
    for i in range(6):
        t = export_date + relativedelta(months=i)
        month_ls.append({'year': t.year, 'month': t.month})
    return month_ls


# 现金流监控表每个一级业务中的二级业务所在行, 新增业务时，修改下面的行数
first_service_opt = [
    {
        'name': '标准DRAM',
        'in_start': 3,
        'in_end': 7,
        'out_start': 10,
        'out_end': 14
    },
    {
        'name': '模组',
        'in_start': 21,
        'in_end': 23,
        'out_start': 25,
        'out_end': 27
    },
    {
        'name': '大带宽产品',
        'in_start': 33,
        'in_end': 35,
        'out_start': 37,
        'out_end': 39
    },
    {
        'name': '设计服务',
        'in_start': 44,
        'in_end': 45,
        'out_start': 47,
        'out_end': 48
    },
]


@api_view(['GET'])
def export_month_monitor(request):
    try:
        export_month = request.GET.get('exportMonth')
        if not export_month:
            return REST_FAIL({'导出时间不能为空'})
        req_user = request.user
        service_filter = get_service_filter(req_user)
        year, month, last_month, qs1, qs2, surplus_data = get_base_data(export_month, service_filter, req_user.username, 'monitor')
        qs = list(qs1) + list(qs2)
        title = '{}年现金流'.format(year)
        template_path = os.path.dirname(__file__) + '/export_template/FBA_Month_Monitor_Template.xlsx'
        if not qs:
            return create_excel_resp(template_path, title)
        query_service_sql = '''select f.id as first_service_id, f.name as first_service, 
                                    s.id as second_service_id , s.name as second_service
                                    from fba_first_service f
                                    left join fba_second_service s on f.id = s.first_service_id
                                    where f.name in ('标准DRAM', '模组', '大带宽产品', '设计服务')
                                    order by f.id,s.id'''
        with connection.cursor() as cursor:
            cursor.execute(query_service_sql)
            service_qs = dictfetchall(cursor)
        df = pd.DataFrame(qs)
        df.rename(columns={'first_service__name': 'first_service', 'second_service__name': 'second_service'},
                  inplace=True)
        ndf = df[['data_type', 'first_service_id', 'first_service', 'second_service_id', 'second_service', 'year',
                  'month', 'in_usd', 'in_cny', 'out_usd', 'out_cny']]
        ndf[['in_usd', 'in_cny', 'out_usd', 'out_cny']] = ndf[['in_usd', 'in_cny', 'out_usd', 'out_cny']].replace({None: 0})
        ndf[['in_usd', 'in_cny', 'out_usd', 'out_cny']] = ndf[['in_usd', 'in_cny', 'out_usd', 'out_cny']].astype(float)
        currency_exchange = float(surplus_data.get('currency_exchange', 0))
        ndf['in'] = round(ndf['in_usd'] * currency_exchange + ndf['in_cny'], 4)
        ndf['out'] = round(ndf['out_usd'] * currency_exchange + ndf['out_cny'], 4)
        ndf.drop(['in_usd', 'in_cny', 'out_usd', 'out_cny'], axis=1, inplace=True)
        moneySum = ndf.groupby(['first_service_id', 'first_service', 'second_service_id', 'second_service',
                                'data_type', 'year', 'month']).sum().reset_index()
        month_ls = create_month_ls(export_month)
        data_type_ls = [{'data_type': 1, 'in': 0, 'out': 0}, {'data_type': 2, 'in': 0, 'out': 0},
                        {'data_type': 3, 'in': 0, 'out': 0}]
        base_data = []
        for m in month_ls:
            for s in service_qs:
                s.update(m)
                for d in data_type_ls:
                    d.update(s)
                    base_data.append(d.copy())
        base_df = pd.DataFrame(base_data)
        base_df = base_df[['first_service_id', 'first_service', 'second_service_id', 'second_service',
                           'data_type', 'year', 'month', 'in', 'out']]
        mdf = pd.concat([base_df, moneySum]).groupby(
            ['first_service_id', 'first_service', 'second_service_id', 'second_service', 'data_type', 'year',
             'month']).sum().reset_index()
        mdf[['in', 'out']] = mdf[['in', 'out']].replace({0: None})
        service_df = pd.DataFrame(service_qs)
        wb = load_workbook(template_path)
        ws = wb['现金流监控总表']
        ws['A1'] = title

        # 循环遍历将数据写入excel
        for item in first_service_opt:
            first_service = item['name']
            fdf = mdf.query('first_service == "{}"'.format(first_service))
            second_service_ls = service_df.query('first_service == "{}"'.format(first_service))['second_service'].tolist()
            j_ls = list(range(item['in_start'], item['in_end'] + 1))
            z_ls = list(range(item['out_start'], item['out_end'] + 1))
            a = 0
            for j, z in zip(j_ls, z_ls):
                second_service = second_service_ls[a]
                sdf = fdf.query('second_service == "{}"'.format(second_service))
                ws.cell(row=j, column=3).value = second_service
                ws.cell(row=z, column=3).value = second_service
                write_col = 5
                for m in month_ls:
                    datas = sdf[sdf.month == m['month']].to_dict('records')
                    ws.cell(row=1, column=write_col).value = '{}年{}月'.format(m['year'], m['month'])

                    for data in datas:
                        ws.cell(row=j, column=write_col).value = data['in']
                        ws.cell(row=z, column=write_col).value = data['out']
                        write_col += 1
                a += 1
        save_path = get_file_path('monitor', 'export_files')
        wb.save(save_path)
        return create_excel_resp(save_path, title)
    except Exception as e:
        logger.error('导出失败, error: {}'.format(traceback.format_exc()))
        return REST_FAIL({'msg': '导出失败, error: {}'.format(str(e))})
